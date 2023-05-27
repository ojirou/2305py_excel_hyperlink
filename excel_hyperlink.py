import requests
import bs4
import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
import subprocess
def main():
    Dir=r'C:\\Users\\user\\python\\web_scraping\\GoogleSearch\\'
#    search_keyword= input('キーワード　>> ') 
    search_keyword='2022年 情報セキュリティ事件 ランキング'
    search_keyword2=search_keyword.replace(" ", "_")
    ExcelName=Dir+search_keyword2+'.xlsx'
    df=pd.read_excel(ExcelName, sheet_name='Sheet1')
    name, ext=os.path.splitext(ExcelName)
    wb = load_workbook(ExcelName)
    ws = wb.active
    column = ws['C']
    for cell in column:
        # cell.value = str(cell.value)
        cell.value = "'" + str(cell.value)
    ws.column_dimensions['B'].width=60
    ws.column_dimensions['C'].width=100
    column_num=3
    row_nums=len(df)
    for row_num in range(1, row_nums+2):
        target_cell = ws.cell(row=row_num, column=column_num) 
        Address=target_cell.value
        target_cell.value = '=HYPERLINK("'+Address+'", "'+Address+'")' 
        target_cell.font  = Font(size=9, color=Color(rgb=None, indexed=None, auto=None, theme=10, tint=0.0, type="theme"))
    FileNameR=name+'_r.xlsx'
    wb.save(FileNameR)
#    os.remove(FileNameR)
    subprocess.Popen(['start', FileNameR], shell=True)
    os.startfile(Dir, operation='open')
if __name__ == "__main__":
    main()